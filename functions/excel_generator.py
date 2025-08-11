"""Module for generating Excel reports."""
from datetime import datetime
from collections import defaultdict
from typing import Dict, Any
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.chart import PieChart, Reference
from openpyxl.chart.label import DataLabelList

from utils import format_month_year_es, parse_date_str, _lighten_color

# Colores para encabezados y totales (más suaves)
HEADER_TOTAL_COLOR = "D9D9D9"

# --- Constantes para Claves de Diccionarios ---
# Claves para el diccionario de datos de entrada 'data'
KEY_CODIGO_CLIENTE = 'codigoCliente'
KEY_RUC = 'ruc'
KEY_RAZON_SOCIAL = 'razonSocial'
KEY_LINEA = 'linea'
KEY_PEDIDO = 'pedido'
KEY_MONTO_ORIGINAL = 'montoOriginal'
KEY_FECHAS_ORDENADAS = 'fechasOrdenadas'
KEY_IS_RESTORED = 'isRestored'
KEY_RESUMEN_MENSUAL = 'resumenMensual'
KEY_MONTOS_ASIGNADOS = 'montosAsignados'

# Claves para el diccionario de estilos 'styles'
STYLE_THIN_BORDER = 'thin_border'
STYLE_MAIN_TITLE_FONT = 'main_title_font'
STYLE_SECTION_TITLE_FONT = 'section_title_font'
STYLE_BOLD_FONT = 'bold_font'
STYLE_ITALIC_FONT = 'italic_font'
STYLE_TABLE_HEADER_FONT = 'table_header_font'
STYLE_CURRENCY_FORMAT = 'currency_format'
STYLE_PERCENTAGE_FORMAT = 'percentage_format'
STYLE_MAIN_COLOR_FILL = 'main_color_fill'
STYLE_LIGHT_MAIN_COLOR_FILL = 'light_main_color_fill'
STYLE_LIGHT_GRAY_FILL = 'light_gray_fill'
STYLE_CENTER_ALIGN = 'center_align'

# --- Funciones de Ayuda para Estilos y Secciones ---
def _define_styles(color_hex: str) -> Dict[str, Any]:
    """Define and return a dictionary of all styles used in the report."""
    white_text = "FFFFFF"
    dark_text = "404040"
    gray_text = "595959"

    return {
        STYLE_THIN_BORDER: Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')),
        STYLE_MAIN_TITLE_FONT: Font(bold=True, size=16, color=white_text),
        STYLE_SECTION_TITLE_FONT: Font(bold=True, size=12, color=white_text),
        STYLE_BOLD_FONT: Font(bold=True),
        STYLE_ITALIC_FONT: Font(italic=True, color=gray_text),
        STYLE_TABLE_HEADER_FONT: Font(bold=True, color=dark_text),
        STYLE_CURRENCY_FORMAT: 'S/ #,##0.00',
        STYLE_PERCENTAGE_FORMAT: '0.00%',
        STYLE_MAIN_COLOR_FILL: PatternFill(start_color=color_hex, end_color=color_hex, fill_type="solid"),
        STYLE_LIGHT_MAIN_COLOR_FILL: PatternFill(start_color=_lighten_color(color_hex, 0.7), end_color=_lighten_color(color_hex, 0.7), fill_type="solid"),
        STYLE_LIGHT_GRAY_FILL: PatternFill(start_color=HEADER_TOTAL_COLOR, end_color=HEADER_TOTAL_COLOR, fill_type="solid"),
        STYLE_CENTER_ALIGN: Alignment(horizontal='center', vertical='center')
    }

def _add_main_title(ws: Worksheet, styles: Dict[str, Any]):
    """Adds the main title to the worksheet."""
    ws.merge_cells('A1:C1')
    cell = ws['A1']
    cell.value = "DISTRIBUCION DE MONTOS POR FECHA"
    cell.font = styles[STYLE_MAIN_TITLE_FONT]
    cell.fill = styles[STYLE_MAIN_COLOR_FILL]
    cell.alignment = styles[STYLE_CENTER_ALIGN]
    ws.row_dimensions[1].height = 25

def _add_info_section(ws: Worksheet, data: Dict[str, Any], start_row: int, styles: Dict[str, Any]) -> int:
    """Adds the 'Información General' section and returns the next available row."""
    current_row = start_row
    ws.merge_cells(f'A{current_row}:B{current_row}')
    cell = ws[f'A{current_row}']
    cell.value = "Información General"
    cell.font = styles[STYLE_SECTION_TITLE_FONT]
    cell.fill = styles[STYLE_MAIN_COLOR_FILL]
    cell.alignment = styles[STYLE_CENTER_ALIGN]
    current_row += 1

    info_data = [
        ("Cód. Cliente:", data.get(KEY_CODIGO_CLIENTE, '')),
        ("RUC:", data.get(KEY_RUC, '')),
        ("Cliente:", data.get(KEY_RAZON_SOCIAL, '')),
        ("Línea:", data.get(KEY_LINEA, '')),
        ("Cód. Pedido:", data.get(KEY_PEDIDO, '')),
        ("Monto Total:", data.get(KEY_MONTO_ORIGINAL, 0)),
        ("Total Letras:", len(data.get(KEY_FECHAS_ORDENADAS, [])))
    ]
    if data.get(KEY_IS_RESTORED):
        info_data.insert(0, ("Origen:", "Restaurado desde respaldo"))

    for label, value in info_data:
        ws[f'A{current_row}'].value = label
        ws[f'B{current_row}'].value = value
        ws[f'A{current_row}'].font = styles[STYLE_BOLD_FONT]
        ws[f'A{current_row}'].border = styles[STYLE_THIN_BORDER]
        ws[f'B{current_row}'].border = styles[STYLE_THIN_BORDER]
        if label == "Monto Total:":
            ws[f'B{current_row}'].number_format = styles[STYLE_CURRENCY_FORMAT]
        if label == "Origen:":
            ws[f'A{current_row}'].font = styles[STYLE_ITALIC_FONT]
            ws[f'B{current_row}'].font = styles[STYLE_ITALIC_FONT]
        current_row += 1
    return current_row

def _add_summary_table(ws: Worksheet, data: Dict[str, Any], start_row: int, styles: Dict[str, Any]) -> tuple:
    """Adds the 'Resumen Mensual' table and returns the end row and chart data references."""
    current_row = start_row
    ws.merge_cells(f'A{current_row}:C{current_row}')
    cell = ws[f'A{current_row}']
    cell.value = "Resumen Mensual"
    cell.font = styles[STYLE_SECTION_TITLE_FONT]
    cell.fill = styles[STYLE_MAIN_COLOR_FILL]
    cell.alignment = styles[STYLE_CENTER_ALIGN]
    current_row += 1

    resumen_mensual = data.get(KEY_RESUMEN_MENSUAL, {})
    total_monto_original = data.get(KEY_MONTO_ORIGINAL, 0)

    headers = ["Mes", "Monto (S/)", "Porcentaje"]
    for col_idx, header_text in enumerate(headers, 1):
        cell = ws.cell(row=current_row, column=col_idx, value=header_text)
        cell.font = styles[STYLE_TABLE_HEADER_FONT]
        cell.fill = styles[STYLE_LIGHT_MAIN_COLOR_FILL]
        cell.border = styles[STYLE_THIN_BORDER]
        cell.alignment = styles[STYLE_CENTER_ALIGN]
    current_row += 1

    sorted_months = sorted(resumen_mensual.keys(), key=lambda x: datetime.strptime(x, "%Y-%m"))

    summary_data_start_row = current_row
    for mes_key in sorted_months:
        monto_mes = resumen_mensual[mes_key]
        porcentaje = (monto_mes / total_monto_original) if total_monto_original > 0 else 0
        
        ws.cell(row=current_row, column=1, value=format_month_year_es(datetime.strptime(mes_key, "%Y-%m"))).border = styles[STYLE_THIN_BORDER]
        ws.cell(row=current_row, column=2, value=monto_mes).number_format = styles[STYLE_CURRENCY_FORMAT]
        ws.cell(row=current_row, column=2).border = styles[STYLE_THIN_BORDER]
        ws.cell(row=current_row, column=3, value=porcentaje).number_format = styles[STYLE_PERCENTAGE_FORMAT]
        ws.cell(row=current_row, column=3).border = styles[STYLE_THIN_BORDER]
        current_row += 1
    summary_data_end_row = current_row - 1

    ws.cell(row=current_row, column=1, value="Totales").font = styles[STYLE_BOLD_FONT]
    ws.cell(row=current_row, column=1).border = styles[STYLE_THIN_BORDER]
    ws.cell(row=current_row, column=1).fill = styles[STYLE_LIGHT_GRAY_FILL]
    
    sum_formula_monto = (
        f'=SUM(B{summary_data_start_row}:B{summary_data_end_row})'
        if summary_data_start_row <= summary_data_end_row
        else 0
    )
    ws.cell(row=current_row, column=2, value=sum_formula_monto).number_format = styles[STYLE_CURRENCY_FORMAT]
    ws.cell(row=current_row, column=2).font = styles[STYLE_BOLD_FONT]
    ws.cell(row=current_row, column=2).border = styles[STYLE_THIN_BORDER]
    ws.cell(row=current_row, column=2).fill = styles[STYLE_LIGHT_GRAY_FILL]

    ws.cell(row=current_row, column=3, value=1).number_format = styles[STYLE_PERCENTAGE_FORMAT]
    ws.cell(row=current_row, column=3).font = styles[STYLE_BOLD_FONT]
    ws.cell(row=current_row, column=3).border = styles[STYLE_THIN_BORDER]
    ws.cell(row=current_row, column=3).fill = styles[STYLE_LIGHT_GRAY_FILL]
    
    return current_row, summary_data_start_row, summary_data_end_row

def _add_pie_chart(ws: Worksheet, start_row: int, end_row: int, anchor: str = 'E3'):
    """Adds a pie chart to the worksheet based on summary data."""
    if start_row > end_row:
        return

    pie_chart = PieChart()
    pie_chart.title = "Distribución Porcentual Mensual"
    data_ref = Reference(ws, min_col=3, min_row=start_row, max_row=end_row)
    labels_ref = Reference(ws, min_col=1, min_row=start_row, max_row=end_row)
    pie_chart.add_data(data_ref, titles_from_data=False)
    pie_chart.set_categories(labels_ref)
    pie_chart.dLbls = DataLabelList(showVal=False, showPercent=True, showCatName=True, showLeaderLines=True)
    pie_chart.width = 12.5
    pie_chart.height = 7.5
    ws.add_chart(pie_chart, anchor)

def _add_detail_table(ws: Worksheet, data: Dict[str, Any], start_row: int, styles: Dict[str, Any]) -> int:
    """Adds the horizontal 'Detalle por Mes' table and returns the next available row."""
    current_row = start_row
    ws.merge_cells(f'A{current_row}:D{current_row}')
    cell = ws[f'A{current_row}']
    cell.value = "Detalle por Mes"
    cell.font = styles[STYLE_SECTION_TITLE_FONT]
    cell.fill = styles[STYLE_MAIN_COLOR_FILL]
    cell.alignment = styles[STYLE_CENTER_ALIGN]
    current_row += 1

    montos_por_mes = defaultdict(list)
    for fecha_str, monto in data.get(KEY_MONTOS_ASIGNADOS, {}).items():
        fecha_obj = parse_date_str(fecha_str)
        montos_por_mes[fecha_obj.strftime("%Y-%m")].append((fecha_obj, monto))

    for mes_key in montos_por_mes:
        montos_por_mes[mes_key].sort(key=lambda item: item[0])

    sorted_months = sorted(montos_por_mes.keys(), key=lambda x: datetime.strptime(x, "%Y-%m"))
    total_monto_original = data.get(KEY_MONTO_ORIGINAL, 0)

    # Headers
    header_row_1, header_row_2 = current_row, current_row + 1
    current_col_idx = 1
    for mes_key in sorted_months:
        mes_obj = datetime.strptime(mes_key, "%Y-%m")
        monto_mes_total = sum(m for _, m in montos_por_mes.get(mes_key, []))
        porcentaje = (monto_mes_total / total_monto_original) if total_monto_original else 0

        # Simplified header creation
        for r, val, fmt, col_offset in [(header_row_1, format_month_year_es(mes_obj), None, 0),
                                        (header_row_1, porcentaje, styles[STYLE_PERCENTAGE_FORMAT], 1),
                                        (header_row_2, "Fechas", None, 0), 
                                        (header_row_2, "Monto (S/)", None, 1)]:
            cell = ws.cell(row=r, column=current_col_idx + col_offset, value=val)
            cell.font = styles[STYLE_TABLE_HEADER_FONT]
            cell.fill = styles[STYLE_LIGHT_MAIN_COLOR_FILL]
            cell.border = styles[STYLE_THIN_BORDER]
            cell.alignment = styles[STYLE_CENTER_ALIGN]
            if fmt: cell.number_format = fmt
        current_col_idx += 2
    current_row += 2

    # Data rows
    max_fechas_por_mes = max((len(v) for v in montos_por_mes.values()), default=0)
    
    for i in range(max_fechas_por_mes):
        current_col_idx = 1
        for mes_key in sorted_months:
            fechas_mes = montos_por_mes[mes_key] # Ya están ordenadas
            if i < len(fechas_mes):
                fecha_obj, monto = fechas_mes[i]
                # Write date
                ws.cell(row=current_row + i, column=current_col_idx, value=fecha_obj.strftime('%d/%m/%Y')).border = styles['thin_border']
                # Write amount
                cell = ws.cell(row=current_row + i, column=current_col_idx + 1, value=monto)
                cell.number_format = styles['currency_format']
                cell.border = styles['thin_border']
            current_col_idx += 2 # Move to the next month's columns

    # Add totals row for the detail section
    detail_total_row = current_row + max_fechas_por_mes
    current_col_idx = 1
    for mes_key in sorted_months:
        monto_mes_total = sum(m for _, m in montos_por_mes.get(mes_key, []))
        ws.cell(row=detail_total_row, column=current_col_idx, value="Total Mes").font = styles['bold_font']
        ws.cell(row=detail_total_row, column=current_col_idx).border = styles['thin_border']
        ws.cell(row=detail_total_row, column=current_col_idx).fill = styles['light_gray_fill']
        cell = ws.cell(row=detail_total_row, column=current_col_idx + 1, value=monto_mes_total)
        cell.number_format = styles['currency_format']
        cell.font = styles['bold_font']
        cell.border = styles['thin_border']
        cell.fill = styles['light_gray_fill']
        current_col_idx += 2
    
    return detail_total_row + 1

def _adjust_column_widths(ws):
    """Sets a fixed width for all columns in the worksheet."""
    # Según el requerimiento, se establece un ancho fijo de 17 para todas las columnas
    # para mantener un diseño consistente, en lugar de un auto-ajuste dinámico.
    for col in ws.columns:
        column_letter = None
        for cell in col:
            if hasattr(cell, 'column_letter'):
                column_letter = cell.column_letter
                break
        if not column_letter:
            continue
        ws.column_dimensions[column_letter].width = 17

# --- Funciones Principales de Creación de Hojas ---
def create_full_report_sheet(wb: Workbook, data: dict, color_hex: str):
    """Crea la hoja principal del reporte orquestando las sub-secciones."""
    ws = wb.active
    ws.title = "Reporte Dashboard"
    
    styles = _define_styles(color_hex)
    
    _add_main_title(ws, styles)
    
    current_row = 3
    current_row = _add_info_section(ws, data, current_row, styles)
    current_row += 2
    
    summary_total_row, summary_start, summary_end = _add_summary_table(ws, data, current_row, styles)
    
    if data.get('resumenMensual'):
        _add_pie_chart(ws, summary_start, summary_end, anchor='E3')
        
    current_row = summary_total_row + 2
    _add_detail_table(ws, data, current_row, styles)
    
    _adjust_column_widths(ws)


def create_payment_detail_sheet(wb: Workbook, data: dict, color_hex: str):
    """Crea una hoja de cálculo con el detalle de cada pago."""
    ws = wb.create_sheet(title="Detalle de Pagos")
    styles = _define_styles(color_hex) # Reutilizar estilos

    # --- Encabezado ---
    ws.merge_cells('A1:C1')
    title_cell = ws['A1']
    title_cell.value = "DETALLE DE VENCIMIENTOS"
    title_cell.font = styles['section_title_font']
    title_cell.fill = styles['main_color_fill']
    title_cell.alignment = styles['center_align']
    ws.row_dimensions[1].height = 20

    # --- Cabeceras de la tabla ---
    headers = ["N°", "Fecha de Vencimiento", "Monto (S/)"]
    for col_idx, header_text in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_idx, value=header_text)
        cell.font = styles['bold_font']
        cell.border = styles['thin_border']
        cell.alignment = styles['center_align']

    # --- Datos ---
    fechas_ordenadas = data.get('fechasOrdenadas', [])
    montos_asignados = data.get('montosAsignados', {})
    start_data_row = 4
    for i, fecha_str in enumerate(fechas_ordenadas):
        monto = montos_asignados.get(fecha_str, 0)
        current_row = start_data_row + i
        ws.cell(row=current_row, column=1, value=i + 1).border = styles['thin_border']
        ws.cell(row=current_row, column=2, value=fecha_str).border = styles['thin_border']
        cell_monto = ws.cell(row=current_row, column=3, value=monto)
        cell_monto.number_format = styles['currency_format']
        cell_monto.border = styles['thin_border']

    # --- Fila de Totales ---
    total_row = start_data_row + len(fechas_ordenadas)
    ws.merge_cells(f'A{total_row}:B{total_row}')
    total_label_cell = ws[f'A{total_row}']
    total_label_cell.value = "Monto Total"
    total_label_cell.font = styles['bold_font']
    total_label_cell.fill = styles['light_gray_fill']
    total_label_cell.border = styles['thin_border']
    total_label_cell.alignment = Alignment(horizontal='right')

    total_value_cell = ws.cell(row=total_row, column=3)
    total_value_cell.value = f"=SUM(C{start_data_row}:C{total_row - 1})"
    total_value_cell.number_format = styles['currency_format']
    total_value_cell.font = styles['bold_font']
    total_value_cell.fill = styles['light_gray_fill']
    total_value_cell.border = styles['thin_border']

    # --- Ajuste de columnas ---
    _adjust_column_widths(ws)